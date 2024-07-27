import openpyxl
from enum import Enum
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet import page
import datetime

session_needs_path = "PlanoReports/SessionNeeds-07-21-2024.xlsx"
session_participants_path = "PlanoReports/SessionsWithParticipants07-21-2024.xlsx"
participant_availabilities_path = "PlanoReports/ParticipantAvailabilities_07-21-2024.xlsx"

header_info = {"A": "Start Time", "B": "Duration", "C": "Title", "D": "Record Session", "E": "Stream Session",
               "F": "Complexity", "G": "Participants", "H": "Notes"}

panel_known_exceptions_list = [
    "Future Worldcons Q&A",
    "Deb Geisler Memorial"
]

large_panel_known_exceptions_list = [
    "African Cultural Influences in Fantasy"
]

class Columns(Enum):
    START = 2
    DURATION = 3
    TITLE = 1
    RECORD = 7
    STREAM = 6
    ADMIN_TAGS = 11
    TECH_NOTES = 13
    ROOM = 4
    FORMAT = 8


# SEC Supported Rooms
class Rooms(Enum):
    CLYDE = "Clyde Auditorium"
    FORTH = "Forth"
    GALA = "Gala"
    H1 = "Hall 1"
    H2 = "Hall 2"
    LOM = "Lomond Auditorium"
    AL1 = "Alsh 1"
    AL2 = "Alsh 2"
    CAR = "Carron"
    D1 = "Dochart 1"
    M1 = "Meeting Academy M1"
    M4 = "Meeting Academy M4"


"""
# All Rooms
class Rooms(Enum):
    CLYDE = "Clyde Auditorium"
    FORTH = "Forth"
    GALA = "Gala"
    H1 = "Hall 1"
    H2 = "Hall 2"
    LOM = "Lomond Auditorium"
    AL1 = "Alsh 1"
    AL2 = "Alsh 2"
    CAR = "Carron"
    D1 = "Dochart 1"
    D2 = "Dochart 2"
    M1 = "Meeting Academy M1"
    M23 = "Meeting Academy M2/3"
    M4 = "Meeting Academy M4"
    AG1 = "Argyll 1"
    AG2 = "Argyll 2"
    AG3 = "Argyll 3"
    C1 = "Castle 1"
    C2 = "Castle 2"
    C3 = "Castle 3"
    SS = "Staffa/Shuna"
    JURA = "Jura"
    BARRA = "Barra"
    INS = "Inspiration"
"""


class Days(Enum):
    THU = "Thursday"  # 8th
    FRI = "Friday"  # 9
    SAT = "Saturday"  # 10
    SUN = "Sunday"  # 11
    MON = "Monday"  # 12


class ConDates(Enum):
    THU = "08"
    FRI = "09"
    SAT = "10"
    SUN = "11"
    MON = "12"


def participant_details():
    participant_details_workbook = openpyxl.load_workbook(participant_availabilities_path)
    participant_details_sheet = participant_details_workbook.active
    last_participant = participant_details_sheet.max_row + 1
    build_details = {}

    for j in range(2, last_participant):
        build_details[participant_details_sheet.cell(row=j, column=2).value] = (
            participant_details_sheet.cell(row=j, column=4).value)

    return build_details


participants = participant_details()


def split_participant_string(participants_string):
    try:
        people = participants_string.split(";")
    except AttributeError:
        people = []
    return people


def modify_attendance_type(attendance_type):
    match attendance_type:
        case None:
            return "UNKNOWN"
        case "hybrid":
            return "in person"
        case _:
            return attendance_type


def session_participants(moderator, session_participants_string):
    people = split_participant_string(session_participants_string)
    moderators = []
    try:
        if moderator.find(";"):
            moderators = split_participant_string(moderator)
    except:
        moderators.append(moderator)
    people_strings = []
    moderator_attendance = "Unknown"
    for mod in moderators:
        try:
            moderator_attendance = participants[mod]
        except KeyError:
            pass
        moderator_attendance = modify_attendance_type(moderator_attendance)
        people_strings.append(str(moderator)+" (Mod, "+moderator_attendance+")")

    for person in people:
        person_attendance = "Unknown"
        try:
            person_attendance = participants[person]
        except KeyError:
            pass
        person_attendance = modify_attendance_type(person_attendance)
        people_strings.append(str(person) + " (" + person_attendance + ")")

    return "\n".join(people_strings)


def session_virtual_participants(moderator, session_participants_string):
    people = split_participant_string(session_participants_string)
    people_strings = []
    moderator_attendance = "Unknown"
    try:
        moderator_attendance = participants[moderator]
    except KeyError:
        pass
    if moderator_attendance is None:
        moderator_attendance = "Unknown"
    virtual_participant_present = moderator_attendance.find("irtu")
    if virtual_participant_present >= 0:
        people_strings.append(str(moderator))

    for person in people:
        person_attendance = "Unknown"
        try:
            person_attendance = participants[person]
        except KeyError:
            pass
        if person_attendance is None:
            person_attendance = "Unknown"
        virtual_participant_present = person_attendance.find("irtu")
        if virtual_participant_present >= 0:
            people_strings.append(str(person))

    if people_strings:
        people = "Virtual: " + "; ".join(people_strings)
    else:
        people = ""
    return people


def session_participant_details():
    session_participants_workbook = openpyxl.load_workbook(session_participants_path)
    session_participants_sheet = session_participants_workbook.active
    last_session = session_participants_sheet.max_row + 1
    session_participant_info = {}

    for k in range(2, last_session):
        session_participant_info[session_participants_sheet.cell(row=k, column=1).value] = (
            session_participants(
                session_participants_sheet.cell(row=k, column=6).value,
                session_participants_sheet.cell(row=k, column=7).value
            ))

    return session_participant_info


def session_virtual_participant_details():
    session_participants_workbook = openpyxl.load_workbook(session_participants_path)
    session_participants_sheet = session_participants_workbook.active
    last_session = session_participants_sheet.max_row + 1
    session_participant_info = {}

    for k in range(2, last_session):
        session_participant_info[session_participants_sheet.cell(row=k, column=1).value] = (
            session_virtual_participants(
                session_participants_sheet.cell(row=k, column=6).value,
                session_participants_sheet.cell(row=k, column=7).value
            ))

    return session_participant_info


sessions_participants = session_participant_details()
virtual_participants = session_virtual_participant_details()


class TechRecord:
    def __init__(self, current_room, day):
        self.room = current_room
        self.day = day
        self.info = []

    def add_info(self, start_time, duration, title, record, stream, admin_tags, notes, interim_room,
                 info_session_participants, info_format):

        if (title.find("CANCELLED") >= 0) or (title.find("WITHDRAWN") >= 0):
            return

        complexity = "Unknown"
        try:
            if "Tech - " in admin_tags:
                offset = admin_tags.find("Tech - ") + 7
                end_tech_complexity = 6
                if admin_tags[offset:].find(";") > 0:
                    end_tech_complexity = admin_tags[offset:].find(";")
                complexity = admin_tags[offset:offset+end_tech_complexity]
        except TypeError:
            pass

        try:
            people_in_session = info_session_participants[title]
        except KeyError:
            people_in_session = ""

        # Corrections and similar, hopefully remove this ASAP
        streaming_rooms = ["Alsh 1", "Alsh 2", "Carron", "Dochart 1", "Meeting Academy M1"]
        if interim_room in streaming_rooms:
            record = "Yes"
            stream = "Yes"

        try:
            virtual_people_in_session = virtual_participants[title]
        except KeyError:
            virtual_people_in_session = ""

        if virtual_people_in_session != "":
            try:
                notes = virtual_people_in_session + "\n" + notes
            except TypeError:
                notes = virtual_people_in_session
            finally:
                if complexity != "AMBER" or complexity != "RED":
                    complexity = "AMBER"

        # Add the session format to the notes
        format_note = ""

        if info_format is None:
            format_given = "None"
        else:
            format_given = info_format

        if title.find("UNAVAIL") >= 0:
            format_given = "Unavailable"

        match format_given:
            case "Rehearsal" | "Ceremony" | "Performance" | "Auction" | "Concert" | "Gameshow" | \
                 "Other" | "Meeting":
                format_note = format_given
            case "Takedown":
                format_note = "Strike"
            case "Setup":
                format_note = "Get-in"
            case "Unavailable":
                format_note = "No public items or participants expected in this session"
            case "Talk":
                if people_in_session.count("\n") == 0:
                    format_note = "Talk given by 1 person"
                else:
                    format_note = "Talk given by " + str(people_in_session.count("\n")+1) + " people"
            case "Panel":
                if (people_in_session.count("\n") == 0) and (len(people_in_session) < 1):
                    if title in panel_known_exceptions_list:
                        format_note = "Panel"
                    else:
                        format_note = "Panel with no people listed against it?"
                        print(title + ": " + format_note)
                else:
                    format_note = "Panel of " + str(people_in_session.count("\n")+1) + " people"
                if (people_in_session.count("\n") >= 5) and (title not in large_panel_known_exceptions_list):
                    print("There are too many participants for " + title + " in " + room)
            case "Interview" | "Dialogue":
                if (people_in_session.count("\n") == 0) and (len(people_in_session) < 1):
                    format_note = "Interview with no people listed against it?"
                    print(title + ": " + format_note)
                else:
                    format_note = "Interview of " + str(people_in_session.count("\n")+1) + " people"
            case "Presentation":
                if (people_in_session.count("\n") == 0) and (len(people_in_session) < 1):
                    format_note = "Presentation with no people listed against it?"
                    print(title + ": " + format_note)
                else:
                    format_note = str(people_in_session.count("\n") + 1) + " presenter(s)"
            case "None":
                print("None can't be done for " + title)
            case _:
                print("Format " + str(info_format) + " not currently covered")

        if format_note != "":
            try:
                notes = format_note + "\n" + notes
            except TypeError:
                notes = format_note

        interim_info = {"A": start_time.strftime("%H:%M"), "B": duration, "C": title,
                        "D": record, "E": stream, "F": complexity, "G": people_in_session, "H": notes}
        self.info.append(interim_info)

    def display_record(self):
        for line in self.info:
            print(line)


def build_records():
    build_tech_records = {}
    for build_room in Rooms:
        for day in Days:
            build_tech_records[str(str(build_room.name) + ":" + str(day.name))] = TechRecord(build_room.value,
                                                                                             day.value)
    return build_tech_records


# print(sessions_participants)
# Ingest the session needs
wb_obj = openpyxl.load_workbook(session_needs_path)

sheet_obj = wb_obj.active
last_row = sheet_obj.max_row + 1

tech_records = build_records()

for i in range(2, last_row):
    line_room = sheet_obj.cell(row=i, column=Columns.ROOM.value).value
    if line_room in Rooms:
        room = Rooms(line_room).name
    else:
        room = "Unknown"
    line_date = sheet_obj.cell(row=i, column=Columns.START.value).value.strftime("%d")
    if line_date in ConDates:
        line_day = ConDates(line_date).name
    else:
        line_day = "Unknown"
    if room == "Unknown" or line_day == "Unknown":
        continue

    tech_records[room+":"+line_day].add_info(
        sheet_obj.cell(row=i, column=Columns.START.value).value,
        sheet_obj.cell(row=i, column=Columns.DURATION.value).value,
        sheet_obj.cell(row=i, column=Columns.TITLE.value).value,
        sheet_obj.cell(row=i, column=Columns.RECORD.value).value,
        sheet_obj.cell(row=i, column=Columns.STREAM.value).value,
        sheet_obj.cell(row=i, column=Columns.ADMIN_TAGS.value).value,
        sheet_obj.cell(row=i, column=Columns.TECH_NOTES.value).value,
        room, sessions_participants,
        sheet_obj.cell(row=i, column=Columns.FORMAT.value).value,
    )

# Build the output report(s)
for sheet_day in Days:
    file_name = "tech_printout - "+sheet_day.value+".xlsx"
    workbook = openpyxl.Workbook()

    workbook.save(filename=file_name)

    for sheet_room in Rooms:
        name = str(str(sheet_room.name) + ":" + str(sheet_day.name))
        title_row = 1
        generator_row = 2
        header_row = 4

        sheet = workbook.create_sheet(sheet_room.name+"-"+sheet_day.name)

        sheet.page_margins = page.PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75)
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = True
        sheet.print_options.gridLines = True

        top_line = ["", Rooms(sheet_room).value, "", "", Days(sheet_day).value]
        sheet.column_dimensions['A'].width = 5.83  # 40 pixels
        sheet.column_dimensions['B'].width = 7.67  # 48 pixels
        sheet.column_dimensions['C'].width = 30  # 245 pixels
        sheet.column_dimensions['D'].width = 7  # 40 pixels
        sheet.column_dimensions['E'].width = 7  # 40 pixels
        sheet.column_dimensions['F'].width = 9.43  # 59 pixels
        sheet.column_dimensions['G'].width = 30  # 312 pixels
        sheet.column_dimensions['H'].width = 30  # 312 pixels
        sheet.row_dimensions[title_row].height = 32
        sheet.row_dimensions[header_row].height = 32

        sheet.append(top_line)
        sheet.merge_cells("B"+str(title_row)+":C"+str(title_row))
        sheet.merge_cells("E"+str(title_row)+":G"+str(title_row))
        sheet['B'+str(title_row)].font = Font(size=24)
        sheet['E'+str(title_row)].font = Font(size=24)

        sheet.append(["Generated: " + datetime.datetime.now().strftime("%H:%M:%S, %d %B %Y") +
                      " from " + session_needs_path])
        # sheet.merge_cells("A"+str(generator_row)+":B"+str(generator_row))

        sheet.append([""])
        sheet.append(header_info)
        cells = ["A"+str(header_row), "D"+str(header_row), "E"+str(header_row)]
        for cell in cells:
            sheet[cell].alignment = Alignment(wrapText=True)

        row_counter = header_row + 1
        for i in range(0, len(tech_records[name].info)):
            sheet.append(tech_records[name].info[i])
            sheet["C" + str(row_counter)].alignment = Alignment(wrapText=True)
            sheet["G" + str(row_counter)].alignment = Alignment(wrapText=True)
            sheet["H" + str(row_counter)].alignment = Alignment(wrapText=True)
            row_counter = row_counter + 1

        workbook.save(file_name)

    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    workbook.save(file_name)
